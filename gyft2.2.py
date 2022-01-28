from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
import pandas as pd
import argparse
import getpass
import re
import json
from bs4 import BeautifulSoup as bs
import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
# Parsing from commmand line
parser = argparse.ArgumentParser()
parser.add_argument("-u", "--user", help="ERP Username/Login ID")
args = parser.parse_args()
if args.user is None:
    args.user = input("Enter you Roll Number: ")
erp_password = getpass.getpass("Enter your ERP password: ")

# Parsing ends

ERP_HOMEPAGE_URL = 'https://erp.iitkgp.ac.in/IIT_ERP3/'
ERP_LOGIN_URL = 'https://erp.iitkgp.ac.in/SSOAdministration/auth.htm'
ERP_SECRET_QUESTION_URL = 'https://erp.iitkgp.ac.in/SSOAdministration/getSecurityQues.htm'


headers = {
    'timeout': '20',
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/51.0.2704.79 Chrome/51.0.2704.79 Safari/537.36',
}

s = requests.Session()
r = s.get(ERP_HOMEPAGE_URL)
soup = bs(r.text, 'html.parser')
sessionToken = soup.find_all(id='sessionToken')[0].attrs['value']
r = s.post(ERP_SECRET_QUESTION_URL, data={'user_id': args.user},
           headers=headers)
secret_question = r.text
print("Your secret question: " + secret_question)
secret_answer = getpass.getpass("Enter the answer to the security question: ")
login_details = {
    'user_id': args.user,
    'password': erp_password,
    'answer': secret_answer,
    'sessionToken': sessionToken,
    'requestedUrl': 'https://erp.iitkgp.ac.in/IIT_ERP3',
}
r = s.post(ERP_LOGIN_URL, data=login_details,
           headers=headers)
try:
    ssoToken = re.search(r'\?ssoToken=(.+)$',r.history[1].headers['Location']).group(1)
except IndexError:
    print("Error: Please make sure the entered credentials are correct!")

timetable_details = {
    'ssoToken': ssoToken,
    'module_id': '16',
    'menu_id': '40',
}
ERP_TIMETABLE_URL = "https://erp.iitkgp.ac.in/Acad/student/view_stud_time_table.jsp"

# This is just a hack to get cookies. TODO: do the standard thing here
abc = s.post('https://erp.iitkgp.ac.in/Acad/student/view_stud_time_table.jsp',
             headers=headers, data=timetable_details)
cookie_val = None
for a in s.cookies:
    if (a.path == "/Acad/"):
        cookie_val = a.value

cookie = {
    'JSESSIONID': cookie_val,
}
r = s.post('https://erp.iitkgp.ac.in/Acad/student/view_stud_time_table.jsp',
           cookies=cookie, headers=headers, data=timetable_details)

table_MN = pd.read_html(r.text)
df = table_MN[2]
wb = Workbook()
df.to_excel(excel_writer="./timetable.xlsx",
            sheet_name='timetable', header=False, index=False)
wb = load_workbook(filename='timetable.xlsx')
ws = wb.active

for i in range(1, 7):
    ws.row_dimensions[i].height = 40

ws.column_dimensions['A'].width = 7
for i in range(2, 11):
    ws.column_dimensions[get_column_letter(i)].width = 20

ft1 = Font(name='ubuntu',
           size=12,
           bold=False,
           italic=False,
           vertAlign=None,
           underline='none',
           strike=False,
           color='00000000')
ft2 = Font(name='Roboto',
           size=15,
           bold=False,
           italic=False,
           vertAlign=None,
           underline='none',
           strike=False,
           color='000000FF')
bd = Border(left=Side(border_style='medium', color='00000000'),
            right=Side(border_style='medium', color='00000000'),
            top=Side(border_style='medium', color='00000000'),
            bottom=Side(border_style='medium', color='00000000'),
            diagonal=Side(border_style='medium', color='00000000'),
            diagonal_direction=0,
            outline=Side(border_style='medium', color='00000000'),
            vertical=Side(border_style='medium', color='00000000'),
            horizontal=Side(border_style='medium', color='00000000')
            )
grayfill = PatternFill(start_color='00D9D9D9',
                       end_color='00D9D9D9',
                       fill_type='solid')
align = Alignment(horizontal='center',
                  vertical='center',
                  text_rotation=0,
                  wrap_text=True,
                  shrink_to_fit=False,
                  indent=0)

for i in range(1, 7):
    for j in range(1, 11):
        cell = ws.cell(row=i, column=j)
        cell.font = ft1
        cell.border = bd
        cell.alignment = align
for cell in ws["1:1"]:
    cell.font = ft2
    cell.fill = grayfill
for cell in ws["A:A"]:
    cell.font = ft2
    cell.fill = grayfill

ws['A1'].value = None

for i in range(2, 7):
    # print('\n')
    for j in range(2, 11):
        cell = ws.cell(row=i, column=j)
        if cell.value != None:
            data = cell.value
            cell.value = data[slice(7)]
            # print(cell.value, end=" ")
wb.save('timetable.xlsx')
