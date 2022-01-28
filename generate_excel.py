
import os
import json
import openpyxl
from openpyxl import load_workbook
INPUT_FILENAME =  "data.txt"
if not os.path.exists(INPUT_FILENAME):
    print ("Input file",INPUT_FILENAME,"does not exist.")
    os._exit(1)


# time slots from numbers
hours = {}
hours[1] = "8:0:AM-8:55:AM"
hours[2] = "9:0:AM-9:55:AM"
hours[3] = "10:0:AM-10:55:AM"
hours[4] = "11:0:AM-11:55:AM"
hours[5] = "12:0:PM-12:55:PM"
hours[6] = "1:0:PM-1:55:PM"
hours[7] = "2:0:PM-2:55:PM"
hours[8] = "3:0:PM-3:55:PM"
hours[9] = "4:0:PM-4:55:PM"
hours[10] = "5:0:PM-5:55:PM"


from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# %%
# Get your timetable
with open(INPUT_FILENAME) as data_file:
        data = json.load(data_file)


#PRINT days in column A of sheet
for i,day in zip(range(2,8),data):
    ws.cell(row=i,column=1).value = day[slice(3)].upper()
wb.save('timetable.xlsx')



#PRINT hours in row 1 of sheet
for i,j in zip(range(2,12),range(1,11)):
    ws.cell(row=1,column=i).value = hours[j]
wb.save('timetable.xlsx')


#ROW NO BY DAY
rowno = {}
rowno["Monday"] = 2
rowno["Tuesday"] = 3
rowno["Wednesday"] = 4
rowno["Thursday"] = 5
rowno["Friday"] = 6
rowno["Saturday"] = 7
# print(rowno["Friday"])


# COLUMN NO BY TIME SLOT
columnno = {}
columnno["8:0:AM-8:55:AM"] = 2
columnno["9:0:AM-9:55:AM"] = 3
columnno["10:0:AM-10:55:AM"] = 4
columnno["11:0:AM-11:55:AM"] = 5
columnno["12:0:PM-12:55:PM"] = 6
columnno["2:0:PM-2:55:PM"] = 8
columnno["5:0:PM-5:55:PM"] = 11
# print(columnno["2:0:PM-2:55:PM"])


# Get subjects code and their respective name from file 'subjects.json'
with open('subjects.json') as data_file:
    subjects = json.load(data_file)
# Find the name of this course
# Use subject name if available, else return subject code
def subject_name(subject_code):
    if (subject_code in subjects.keys()):
                return subjects[subject_code].title()
    else:
        return subject_code
# print(subjects['CH21206'].title())                


for day in data:
    for time in data[day]:
        cell = ws.cell(row = rowno[day],column = columnno[time])
        cell.value = subject_name(data[day][time][0])
        if(data[day][time][2]>1):
            ws.merge_cells(start_row=rowno[day],start_column=columnno[time],end_row = rowno[day],end_column=columnno[time]+data[day][time][2]-1)
wb.save('timetable.xlsx')



from openpyxl.utils import get_column_letter

ws.row_dimensions[1].width = 40
for i in range (2,7):
    ws.row_dimensions[i].height = 34

ws.column_dimensions['A'].width = 8
for i in range (2,12):
    ws.column_dimensions[get_column_letter(i)].width = 17

wb.save('timetable.xlsx')


from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

ft1 = Font(name='ubuntu',
           size=12,
           bold=False,
           italic=False,
           vertAlign=None,
           underline='none',
           strike=False,
           color='00000000')
ft2 = Font(name='Roboto',
           size=15,
           bold=False,
           italic=False,
           vertAlign=None,
           underline='none',
           strike=False,
           color='000000FF')
bd = Border(left=Side(border_style='medium', color='00000000'),
            right=Side(border_style='medium', color='00000000'),
            top=Side(border_style='medium', color='00000000'),
            bottom=Side(border_style='medium', color='00000000'),
            diagonal=Side(border_style='medium', color='00000000'),
            diagonal_direction=0,
            outline=Side(border_style='medium', color='00000000'),
            vertical=Side(border_style='medium', color='00000000'),
            horizontal=Side(border_style='medium', color='00000000')
            )
grayfill = PatternFill(start_color='00D9D9D9',
                       end_color='00D9D9D9',
                       fill_type='solid')
align = Alignment(horizontal='center',
                  vertical='center',
                  text_rotation=0,
                  wrap_text=True,
                  shrink_to_fit=False,
                  indent=0)



#FIRST ROW AND COLUMN STYLES
for i in range(1, 7):
    for j in range(1, 12):
        cell = ws.cell(row=i, column=j)
        cell.font = ft1
        cell.border = bd
        cell.alignment = align
for cell in ws["1:1"]:
    cell.font = ft2
    cell.fill = grayfill
for cell in ws["A:A"]:
    cell.font = ft2
    cell.fill = grayfill

ws['A1'].value = None
wb.save('timetable.xlsx')



